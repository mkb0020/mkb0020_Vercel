from flask import Blueprint, request, send_file, jsonify
import io
import os
import tempfile
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.drawing.image import Image

truedelta_bp = Blueprint('truedelta', __name__)

# ----------------------------  DETAIL TAB HEADERS ----------------------------
DETAIL_TAB_HEADERS = [
    "SKU", "PRODUCT DESCRIPTION", "PACKAGE", "TIER", "LICENSE SHIFT ELIGIBLE",
    "ALLOCATION STATUS", "EXISTING QTY", "NEW QTY", "(+)LICENSE SHIFT",
    "(-)LICENSE SHIFT", "DELTA QTY", "UNIT LIST PRICE", "DISCOUNT OFF LIST",
    "UNIT NET PRICE", "EXISTNG NET PRICE (months)", "NEW NET PRICE (months)",
    "EXISTING PRORATED PRICE", "NEW PRORATED PRICE", "PROJECTED CREDIT",
    "PROJECTED CHARGE", "TRUE DELTA NET COST"
]

SOURCE_HEADER_ALIASES = {
    "SKU|ProductID": "SKU",
    "PRODUCT|Product|ProductDescription|ItemDescription": "PRODUCT DESCRIPTION",
    "PACKAGE NAME|PackageName|Suite|SUITE|SuiteName": "PACKAGE",
    "Tier|TIER|CommitStatus|Commit|FullCommit": "TIER",
    "License Shift Eligible|LICENSE SHIFT ELIGIBLE|UsageCategory|USAGE CATEGORY|Usage Category|Usage": "LICENSE SHIFT ELIGIBLE",
    "Original Purchase|Initial Quantity|InitialQTY|ExistingQuantity|ExistingQTY|Existing Quantity|Exisiting QTY": "EXISTING QTY",
    "Actual Deployment|New Quantity|NewQTY|NewQuantity|New Quantity| New QTY": "NEW QTY",
    "License Shift Additional QTY|Added For Value Shift|ADDED FOR VALUE SHIFT|Added for License Shift|Positive License Shift": "(+)LICENSE SHIFT",
    "License Shift  Returned QTY|Returned For Value Shift|RETURNED FOR VALUE SHIFT|Returned for License Shift|Negaiv License Shift": "(-)LICENSE SHIFT",
    "Unit List Price|List Price|List": "UNIT LIST PRICE"
}

DROP_COLUMNS = {
    "Delta", "Annual True Delta", "True Delta Cost", "Actual Usage Cost", "Original Cost",
    "Delta Cost", "Adjustment", "Distributor Net Cost", "Discount (%)",
    "Allocation Exceeded", "Total Duration (Months)", "Remaining Duration (Months)"
}


# ----------------------------  HELPER: SHEET → LIST OF DICTS ----------------------------
def _sheet_to_dicts(ws):
    """Read an openpyxl worksheet into a list of dicts keyed by the header row."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"__col{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            result.append(dict(zip(headers, row)))
    return result


def _parse_date(value):
    """Try to return a datetime from various input types; return value unchanged on failure."""
    if value is None:
        return value
    if isinstance(value, datetime):
        return value
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d %H:%M:%S")
    except ValueError:
        pass
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    return value


# ----------------------------  PRICING ENGINE ----------------------------
class TrueDeltaPricing:
    def get_discount(self, PricingType, Percent, LineDiscount, UnitCost, UnitList):
        PricingType = PricingType.strip().upper()
        if PricingType == "HOLD BACK":
            return (LineDiscount / 100) - Percent
        elif PricingType == "MARKUP":
            return 1 - ((UnitCost * (Percent + 1)) / UnitList) if UnitList else 0
        elif PricingType == "MARGIN":
            return 1 - ((UnitCost / (1 - Percent)) / UnitList) if UnitList else 0
        return 0

    def get_UnitNP(self, PricingType, Percent, LineDiscount, UnitCost, UnitList):
        PricingType = PricingType.strip().upper()
        if PricingType == "HOLD BACK":
            return UnitList - (UnitList * ((LineDiscount / 100) - Percent))
        elif PricingType == "MARKUP":
            return UnitCost * (Percent + 1)
        elif PricingType == "MARGIN":
            return UnitCost / (1 - Percent) if (1 - Percent) != 0 else UnitList
        return UnitList

    def get_consumption(self, TFqty, ExistingQTY, NewQTY, vsADD, vsRETURNED,
                        STATUS1=None, STATUS2=None, STATUS3=None, STATUS4=None, STATUS5=None, TIER=None):
        STATUS1 = STATUS1 or "Within Allocation"
        STATUS2 = STATUS2 or "Allocation Exceeded"
        STATUS3 = STATUS3 or "(+)License Shift"
        STATUS4 = STATUS4 or "(-)License Shift"
        STATUS5 = STATUS5 or "FIXED"
        if TFqty == 0:
            return STATUS1
        elif TFqty > 0:
            return STATUS3 if vsADD > 0 else STATUS2
        elif TFqty < 0:
            if TIER and TIER.upper() == "FIXED":
                return STATUS5
            elif vsRETURNED > 0:
                return STATUS4
        return "UNKNOWN"

    def get_LineCM(self, TFqty, vsRETURNED, LineExistingProratedNP):
        return 0 if TFqty == 0 else LineExistingProratedNP

    def get_LineINV(self, TFqty, LineNewProratedNP):
        return 0 if TFqty == 0 else LineNewProratedNP


# ----------------------------  DETAILS TAB FORMATTING ----------------------------
class TrueDeltaDetailsFormat:
    def __init__(self, filepath, sheet_name, total_months=None, remaining_months=None):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.total_months = total_months
        self.remaining_months = remaining_months
        self.wb = load_workbook(filepath)
        self.ws = self.wb[sheet_name]
        self.HeaderColor = PatternFill(start_color="4B1395", end_color="4B1395", fill_type="solid")
        self.BodyColor = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.ThinBorders = Side(style="thin", color="000000")
        self.ThickBorder = Side(style="thick", color="000000")
        self.DoubleBorders = Side(border_style="double", color="000000")
        self.HeaderFont = Font(name="Aptos Narrow", bold=True, color="FFFFFF", size=10)
        self.BodyFont = Font(name="Aptos Narrow", bold=False, color="000000", size=10)
        self.Leftys = {"SKU", "PRODUCT DESCRIPTION"}
        self.Middles = {
            "PACKAGE", "TIER", "LICENSE SHIFT ELIGIBLE", "ALLOCATION STATUS",
            "EXISTING QTY", "NEW QTY", "(+)LICENSE SHIFT", "(-)LICENSE SHIFT", "DELTA QTY"
        }
        self.Rightys = {
            "UNIT LIST PRICE", "DISCOUNT OFF LIST", "UNIT NET PRICE",
            "EXISTNG NET PRICE (months)", "NEW NET PRICE (months)",
            "EXISTING PRORATED PRICE", "NEW PRORATED PRICE",
            "PROJECTED CREDIT", "PROJECTED CHARGE", "TRUE DELTA NET COST"
        }
        self.CurrencyColumns = list(self.Rightys)
        self.PercentColumns = ["DISCOUNT OFF LIST"]

    def set_deets_column_widths(self):
        for col in ["A", "E"]:
            self.ws.column_dimensions[col].width = 13
        for col in ["B", "C", "F"]:
            self.ws.column_dimensions[col].width = 20
        for col in ["D", "G", "H", "I", "K", "L", "M", "N"]:
            self.ws.column_dimensions[col].width = 12
        for col in ["O", "P", "Q", "R", "S", "T", "U"]:
            self.ws.column_dimensions[col].width = 15

    def apply_alignment(self):
        header_row = next(self.ws.iter_rows(min_row=1, max_row=1))
        col_index_to_name = {idx + 1: cell.value for idx, cell in enumerate(header_row) if cell.value}
        for cell in header_row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for row in self.ws.iter_rows(min_row=2):
            for idx, cell in enumerate(row, start=1):
                col_name = col_index_to_name.get(idx)
                if col_name in self.Leftys:
                    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                elif col_name in self.Middles:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_name in self.Rightys:
                    cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)

    def apply_number_formats(self):
        header = [cell.value for cell in self.ws[1]]
        for col_name in self.CurrencyColumns:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in self.ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        for col_name in self.PercentColumns:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in self.ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = "0.00%"

    def style_header_row(self):
        for col in range(1, 22):
            cell = self.ws.cell(1, col)
            cell.fill = self.HeaderColor
            cell.font = self.HeaderFont
            cell.border = Border(
                left=self.ThinBorders, right=self.ThinBorders,
                top=self.ThinBorders, bottom=self.DoubleBorders
            )
        self.ws.row_dimensions[1].height = 45

    def style_body_rows(self):
        for row in range(2, self.ws.max_row + 1):
            self.ws.row_dimensions[row].height = 15
            for col in range(1, 22):
                cell = self.ws.cell(row, col)
                cell.font = self.BodyFont
                cell.fill = self.BodyColor
                cell.border = Border(
                    left=self.ThinBorders, right=self.ThinBorders,
                    top=self.ThinBorders, bottom=self.ThinBorders
                )

    def apply_outline_borders(self):
        for row in self.ws.iter_rows(min_row=1, max_row=self.ws.max_row, min_col=1, max_col=21):
            for cell in row:
                borders = {k: getattr(cell.border, k) for k in ("left", "right", "top", "bottom")}
                if cell.row == 1:
                    borders["top"] = self.ThickBorder
                if cell.row == self.ws.max_row:
                    borders["bottom"] = self.ThickBorder
                if cell.column == 1:
                    borders["left"] = self.ThickBorder
                if cell.column == 21:
                    borders["right"] = self.ThickBorder
                cell.border = Border(**borders)

    def freeze_header(self):
        self.ws.freeze_panes = "A2"

    def update_dynamic_headers(self):
        total_lbl     = f"{int(self.total_months)} months" if self.total_months is not None else "? months"
        remaining_lbl = f"{int(self.remaining_months)} months" if self.remaining_months is not None else "? months"
        self.ws["O1"].value = f"EXISTING\nNET PRICE\n({total_lbl})"
        self.ws["P1"].value = f"NEW\nNET PRICE\n({total_lbl})"
        self.ws["Q1"].value = f"EXISTING\nPRORATE PRICE\n({remaining_lbl})"
        self.ws["R1"].value = f"NEW\nPRORATE PRICE\n({remaining_lbl})"
        self.ws["S1"].value = "ESTIMATED\nCREDIT"
        self.ws["T1"].value = "ESTIMATED\nINVOICE"
        self.ws["U1"].value = "TRUE DELTA\nNET COST"
        for col in ["O", "P", "Q", "R"]:
            self.ws[f"{col}1"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def bold_delta_rows(self, start_row, end_row, delta_col="K", max_col=21):
        for row in range(start_row, end_row + 1):
            delta_value = self.ws[f"{delta_col}{row}"].value
            if delta_value and delta_value != 0:
                for col in range(1, max_col + 1):
                    cell = self.ws.cell(row=row, column=col)
                    cell.font = Font(name=cell.font.name, bold=True, color=cell.font.color, size=cell.font.size)

    def format(self):
        self.set_deets_column_widths()
        self.apply_alignment()
        self.apply_number_formats()
        self.style_header_row()
        self.style_body_rows()
        self.apply_outline_borders()
        self.freeze_header()
        self.update_dynamic_headers()
        self.bold_delta_rows(start_row=2, end_row=self.ws.max_row)
        self.wb.save(self.filepath)


# ----------------------------  SUMMARY TAB FORMATTING ----------------------------
class TrueDeltaSummaryFormat:
    def __init__(self, filepath, sheet_name):
        self.filepath = filepath
        self.sheet_name = sheet_name
        self.wb = load_workbook(filepath)
        self.ws = self.wb[sheet_name]
        self.HeaderColor = PatternFill(start_color="4B1395", end_color="4B1395", fill_type="solid")
        self.SubHeaderColor = PatternFill(start_color="CBCEDF", end_color="CBCEDF", fill_type="solid")
        self.BodyColor = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        self.ThinBorders = Side(style="thin", color="000000")
        self.ThickBorder = Side(style="thick", color="000000")
        self.DoubleBorders = Side(border_style="double", color="000000")
        self.MainHeaderFont = Font(name="Aptos Display", bold=True, color="FFFFFF", size=14)
        self.HeaderFont = Font(name="Aptos Narrow", bold=True, color="FFFFFF", size=10)
        self.SubHeaderFont = Font(name="Aptos Narrow", bold=True, color="000000", size=10)
        self.BodyFont = Font(name="Aptos Narrow", bold=False, color="000000", size=10)
        self.CenterAlign = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.LeftAlign = Alignment(horizontal="left", vertical="center", indent=1)
        self.RightAlign = Alignment(horizontal="right", vertical="center", indent=1)
        self.CurrencyColumns = {"UNIT LIST PRICE", "UNIT NET PRICE", "PROJECTED CREDIT", "PROJECTED CHARGE", "TRUE DELTA NET COST"}

    def set_column_widths(self):
        for col in ["A", "B"]:
            self.ws.column_dimensions[col].width = 26
        for col in ["C", "D", "E", "F", "G"]:
            self.ws.column_dimensions[col].width = 11
        for col in ["H", "I", "J"]:
            self.ws.column_dimensions[col].width = 19
        self.ws.column_dimensions["K"].width = 1

    def style_main_header(self):
        cell = self.ws["A1"]
        cell.alignment = self.CenterAlign
        cell.fill = self.HeaderColor
        cell.font = self.MainHeaderFont
        cell.border = Border(left=self.ThickBorder, right=self.ThickBorder, top=self.ThinBorders, bottom=self.DoubleBorders)
        self.ws.row_dimensions[1].height = 22
        self.ws.row_dimensions[2].height = 5

    def style_general_info(self):
        for row in range(3, 20):
            label_cell = self.ws[f"A{row}"]
            value_cell = self.ws[f"B{row}"]
            label_cell.fill = self.SubHeaderColor
            label_cell.font = self.SubHeaderFont
            label_cell.alignment = self.LeftAlign
            label_cell.border = Border(left=self.ThickBorder, right=self.ThinBorders, top=self.ThinBorders, bottom=self.ThinBorders)
            value_cell.fill = self.BodyColor
            value_cell.font = self.BodyFont
            value_cell.alignment = self.RightAlign
            value_cell.border = Border(left=self.ThinBorders, right=self.ThickBorder, top=self.ThinBorders, bottom=self.ThinBorders)
            if row in [10, 11, 12, 13]:
                value_cell.number_format = "DD-MMM-YYYY"
            if row in [17, 18, 19]:
                value_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        self.ws["A3"].border = Border(left=self.ThickBorder, right=self.ThinBorders, top=self.ThickBorder, bottom=self.ThinBorders)
        self.ws["B3"].border = Border(left=self.ThinBorders, right=self.ThickBorder, top=self.ThickBorder, bottom=self.ThinBorders)
        self.ws["A19"].border = Border(left=self.ThickBorder, right=self.ThinBorders, top=self.ThinBorders, bottom=self.ThickBorder)
        self.ws["B19"].border = Border(left=self.ThinBorders, right=self.ThickBorder, top=self.ThinBorders, bottom=self.ThickBorder)
        for row in range(3, 21):
            self.ws.row_dimensions[row].height = 12
        self.ws.row_dimensions[20].height = 5

    def style_items_table(self, start_row, end_row):
        self.Leftys = {"SKU"}
        self.Middles = {"ALLOCATION STATUS", "EXISTING QTY", "NEW QTY", "DELTA QTY", "UNIT LIST PRICE", "UNIT NET PRICE"}
        self.Rightys = {"PROJECTED CREDIT", "PROJECTED CHARGE", "TRUE DELTA NET COST"}
        for row in range(start_row, end_row + 1):
            self.ws.row_dimensions[row].height = 30 if row == start_row else 15
            for col in range(1, 11):
                cell = self.ws.cell(row=row, column=col)
                if row == start_row:
                    cell.fill = self.HeaderColor
                    cell.font = self.HeaderFont
                    cell.alignment = self.CenterAlign
                    cell.border = Border(left=self.ThinBorders, right=self.ThinBorders, top=self.ThinBorders, bottom=self.DoubleBorders)
                else:
                    cell.fill = self.BodyColor
                    cell.font = self.BodyFont
                    header_value = self.ws.cell(row=start_row, column=col).value
                    if header_value in self.Middles:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    elif header_value in self.Rightys:
                        cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
                    elif header_value in self.Leftys:
                        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
                    bottom_style = self.DoubleBorders if row == end_row - 2 else self.ThinBorders
                    cell.border = Border(left=self.ThinBorders, right=self.ThinBorders, top=self.ThinBorders, bottom=bottom_style)

    def style_totals_row(self, row):
        self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        cell = self.ws.cell(row=row, column=1)
        cell.value = "TOTAL:"
        cell.font = self.SubHeaderFont
        cell.alignment = self.LeftAlign
        cell.fill = self.SubHeaderColor
        cell.border = Border(left=self.ThinBorders, right=self.ThinBorders, bottom=self.ThickBorder)
        for col in ["H", "I", "J"]:
            c = self.ws[f"{col}{row}"]
            c.font = self.SubHeaderFont
            c.alignment = self.RightAlign
            c.fill = self.SubHeaderColor
            c.border = Border(left=self.ThinBorders, right=self.ThinBorders, bottom=self.ThickBorder)
        self.ws.row_dimensions[row + 1].height = 5

    def style_notes(self, start_row):
        for i in range(4):
            row = start_row + i
            for col in range(1, 11):
                cell = self.ws.cell(row=row, column=col)
                cell.border = Border(
                    left=self.ThickBorder if col == 1 else self.ThinBorders,
                    right=self.ThickBorder if col == 10 else self.ThinBorders,
                    top=self.ThickBorder if i == 0 else self.ThinBorders,
                    bottom=self.ThickBorder if i == 3 else self.ThinBorders
                )
            self.ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
            anchor = self.ws.cell(row=row, column=1)
            anchor.alignment = self.LeftAlign
            anchor.fill = self.HeaderColor if i == 0 else self.SubHeaderColor
            anchor.font = self.HeaderFont if i == 0 else self.BodyFont

    def apply_outline_borders(self, item_end_row, notes_start_row):
        self.item_end_row = item_end_row
        self.notes_start_row = notes_start_row
        for row in self.ws.iter_rows(min_row=21, max_row=item_end_row + 1, min_col=1, max_col=10):
            for cell in row:
                self._outline(cell)

    def _outline(self, cell):
        last_item_row = self.item_end_row
        totals_row = last_item_row
        notes_start_row = self.notes_start_row
        borders = {
            "left": self.ThickBorder if cell.column == 1 else self.ThinBorders,
            "right": self.ThickBorder if cell.column == 10 else self.ThinBorders,
            "top": self.ThickBorder if cell.row in [3, 21, notes_start_row] else self.ThinBorders,
            "bottom": (
                self.DoubleBorders if cell.row == last_item_row - 1 else
                self.ThickBorder if cell.row == totals_row else
                self.ThinBorders
            )
        }
        cell.border = Border(**borders)

    def apply_summary_outline(self, last_notes_row):
        for row in range(2, last_notes_row + 1):
            for col in range(1, 11):
                cell = self.ws.cell(row=row, column=col)
                borders = {
                    "left": self.ThickBorder if col == 1 else cell.border.left,
                    "right": self.ThickBorder if col == 10 else cell.border.right,
                    "top": self.ThickBorder if row == 1 else cell.border.top,
                    "bottom": self.ThickBorder if row == last_notes_row else cell.border.bottom
                }
                cell.border = Border(**borders)

    def apply_currency_format(self, start_row, end_row):
        header = [self.ws.cell(row=start_row, column=col).value for col in range(1, 11)]
        for col_name in self.CurrencyColumns:
            if col_name in header:
                col_idx = header.index(col_name) + 1
                for row in range(start_row + 1, end_row + 2):
                    cell = self.ws.cell(row=row, column=col_idx)
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    def White_Out(self):
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        for start, end in [("A2", "J2"), ("C3", "J19"), ("A20", "J20")]:
            for row in self.ws[start:end]:
                for cell in row:
                    cell.fill = white_fill

    def bold_delta_rows(self, start_row, end_row, delta_col="J", max_col=10):
        for row in range(start_row, end_row + 1):
            delta_value = self.ws[f"{delta_col}{row}"].value
            if delta_value and delta_value != 0:
                for col in range(1, max_col + 1):
                    cell = self.ws.cell(row=row, column=col)
                    cell.font = Font(name=cell.font.name, bold=True, color=cell.font.color, size=cell.font.size)

    def format(self, item_end_row):
        self.item_end_row = self.ws.max_row - 5
        self.notes_start_row = item_end_row + 3
        self.set_column_widths()
        self.style_main_header()
        self.style_general_info()
        self.style_items_table(start_row=21, end_row=item_end_row)
        self.style_totals_row(item_end_row + 1)
        self.style_notes(item_end_row + 3)
        self.apply_outline_borders(item_end_row, item_end_row + 3)
        self.apply_summary_outline(last_notes_row=self.notes_start_row + 3)
        self.apply_currency_format(start_row=21, end_row=item_end_row)
        self.White_Out()
        self.wb.save(self.filepath)


# ----------------------------  SUMMARY BUILDER ----------------------------
class TrueDeltaSummaryBuilder:
    def __init__(self, filepath, info, summary_values, details_rows, logo_path=None):
        """
        summary_values: list of 10 values read from Summary column B, rows 1–10
                        (maps to the original summary_df.iloc[0..9, 1])
        details_rows:   list of dicts — filtered to DELTA QTY != 0
        """
        self.filepath = filepath
        self.info = info
        self.summary_values = summary_values
        self.details_rows = details_rows
        self.logo_path = logo_path
        self.wb = load_workbook(filepath)
        self.ws = self.wb["Summary"]
        self.main_header = "TRUE DELTA ESTIMATE"
        self.general_headers = [
            "Created By:", "Reseller:", "Reseller Account #:", "End Customer:",
            "Procurement Plan:", "Reference ID:", "Vendor Account:",
            "Agreement Start Date:", "Agreement End Date:", "Report Date:",
            "Next True Δ Date:", "Agreement Duration (months):",
            "Remaining Duration (months):", "Billing Type:",
            "Projected Credit:", "Projected Charge:", "True Delta Net Cost:"
        ]
        self.item_headers = [
            "SKU", "ALLOCATION STATUS", "EXISTING QTY", "NEW QTY", "DELTA QTY",
            "UNIT LIST PRICE", "UNIT NET PRICE", "PROJECTED CREDIT", "PROJECTED CHARGE", "TRUE DELTA NET COST"
        ]

    def clear_sheet(self):
        for row in self.ws.iter_rows():
            for cell in row:
                cell.value = None
                cell._style = None

    def write_main_header(self):
        self.ws.merge_cells("A1:J1")
        self.ws["A1"].value = self.main_header

    def write_general_info(self):
        raw_values = [
            self.info.get("UserName", ""),
            self.info.get("PartnerName", ""),
            self.info.get("PartnerAcct", ""),
            self.summary_values[0],   # End Customer        (row 1, col B)
            self.summary_values[1],   # Procurement Plan    (row 2, col B)
            self.summary_values[2],   # Reference ID        (row 3, col B)
            self.summary_values[3],   # Vendor Account      (row 4, col B)
            self.summary_values[4],   # Agreement Start     (row 5, col B)
            self.summary_values[5],   # Agreement End       (row 6, col B)
            self.summary_values[6],   # Report Date         (row 7, col B)
            self.summary_values[7],   # Next True-Δ Date    (row 8, col B)
            self.summary_values[8],   # Agreement Duration  (row 9, col B)
            self.summary_values[9],   # Remaining Duration  (row 10, col B)
            self.info.get("BillingType", ""),
            0, 0, 0
        ]
        date_rows = {10, 11, 12, 13}  
        for i, (label, value) in enumerate(zip(self.general_headers, raw_values), start=3):
            self.ws[f"A{i}"] = label
            if i in date_rows:
                parsed = _parse_date(value)
                self.ws[f"B{i}"].value = parsed
                if isinstance(parsed, datetime):
                    self.ws[f"B{i}"].number_format = "DD-MMM-YYYY"
            else:
                self.ws[f"B{i}"].value = value

    def build_summary_items(self):
        start_row = 21
        for col_idx, header in enumerate(self.item_headers, start=1):
            self.ws.cell(row=start_row, column=col_idx).value = header
        for i, row in enumerate(self.details_rows):
            r = start_row + 1 + i
            self.ws.cell(row=r, column=1).value = row.get("SKU", "")
            self.ws.cell(row=r, column=2).value = row.get("ALLOCATION STATUS", "")
            self.ws.cell(row=r, column=3).value = row.get("EXISTING QTY", 0)
            self.ws.cell(row=r, column=4).value = row.get("NEW QTY", 0)
            self.ws.cell(row=r, column=5).value = row.get("DELTA QTY", 0)
            self.ws.cell(row=r, column=6).value = row.get("UNIT LIST PRICE", 0)
            self.ws.cell(row=r, column=7).value = row.get("UNIT NET PRICE", 0)
            self.ws.cell(row=r, column=8).value = row.get("PROJECTED CREDIT", 0)
            self.ws.cell(row=r, column=9).value = row.get("PROJECTED CHARGE", 0)
            self.ws.cell(row=r, column=10).value = row.get("TRUE DELTA NET COST", 0)
        return start_row + 1 + len(self.details_rows)

    def write_totals_row(self, row_idx):
        total_credit  = sum(r.get("PROJECTED CREDIT", 0) or 0 for r in self.details_rows)
        total_invoice = sum(r.get("PROJECTED CHARGE", 0) or 0 for r in self.details_rows)
        total_delta   = sum(r.get("TRUE DELTA NET COST", 0) or 0 for r in self.details_rows)
        self.ws.cell(row=row_idx - 1, column=1).value = "TOTAL:"
        self.ws.cell(row=row_idx - 1, column=8).value = total_credit
        self.ws.cell(row=row_idx - 1, column=9).value = total_invoice
        self.ws.cell(row=row_idx - 1, column=10).value = total_delta
        self.ws["B17"].value = total_credit
        self.ws["B18"].value = total_invoice
        self.ws["B19"].value = total_delta

    def write_notes(self, start_row):
        notes = [
            "Notes:",
            "The pricing included in this document is for budgetary purposes only as the quantities are subject to change.",
            "The True Delta Order will automatically book in PsiWave's Portal on the True Delta date unless a manual order is booked prior to the True Delta Date.",
            "This document does NOT include usage/overage fees and applicable taxes. Actual amounts billed may differ based on the final change date, billing model and bill date."
        ]
        for i, note in enumerate(notes):
            self.ws.cell(row=start_row + i - 1, column=1).value = note

    def insert_logo(self, logo_bytes=None):
        try:
            if logo_bytes:
                img = Image(io.BytesIO(logo_bytes))
            elif self.logo_path and os.path.exists(self.logo_path):
                img = Image(self.logo_path)
            else:
                return
            img.width = 250
            img.height = 110
            img.anchor = "I3"
            self.ws.add_image(img)
        except Exception as e:
            print(f"Logo insertion skipped: {e}")

    def build(self, logo_bytes=None):
        self.clear_sheet()
        self.write_main_header()
        self.write_general_info()
        last_item_row = self.build_summary_items()
        self.write_totals_row(last_item_row + 1)
        self.write_notes(last_item_row + 3)
        self.insert_logo(logo_bytes=logo_bytes)
        self.wb.save(self.filepath)


# ----------------------------  COLUMNS PROCESSOR ----------------------------
class TrueDeltaColumns:
    def __init__(self, rows, user_input, pricing_engine, header_aliases, final_headers):
        """rows: list of dicts read from the details sheet via _sheet_to_dicts()"""
        self.rows = [dict(r) for r in rows]
        self.user_input = user_input
        self.pricing_engine = pricing_engine
        self.header_aliases = header_aliases
        self.final_headers = final_headers

    def normalize_headers(self):
        if not self.rows:
            return
        rename_map = {}
        existing_keys = set(self.rows[0].keys())
        for aliases, standard_name in self.header_aliases.items():
            for alias in aliases.split("|"):
                alias = alias.strip()
                if alias in existing_keys:
                    rename_map[alias] = standard_name
        self.rows = [{rename_map.get(k, k): v for k, v in row.items()} for row in self.rows]

    def _process_row(self, row):
        PricingType  = self.user_input["PricingType"]
        Percent      = self.user_input["PercentInput"] / 100

        LineDiscount     = float(row.get("Discount (%)", 0) or 0)
        UnitCost         = float(row.get("Distributor Net Cost", 0) or 0)
        UnitList         = float(row.get("UNIT LIST PRICE", 0) or 0)
        ExistingQTY      = int(row.get("EXISTING QTY", 0) or 0)
        NewQTY           = int(row.get("NEW QTY", 0) or 0)
        vsADD            = row.get("(+)LICENSE SHIFT", 0) or 0
        vsRETURNED       = int(row.get("(-)LICENSE SHIFT", 0) or 0)
        TIER             = row.get("TIER", "") or ""
        EAduration       = float(row.get("Total Duration (Months)", 1) or 1)
        ProratedDuration = float(row.get("Remaining Duration (Months)", 1) or 1)

        TFqty = 0 if vsRETURNED == 0 and ExistingQTY > NewQTY else NewQTY - ExistingQTY

        UnitNP               = self.pricing_engine.get_UnitNP(PricingType, Percent, LineDiscount, UnitCost, UnitList)
        LineExistingNP       = ExistingQTY * UnitNP * EAduration
        LineNewNP            = NewQTY * UnitNP * EAduration
        LineExistingProratedNP = (LineExistingNP / EAduration) * ProratedDuration if EAduration else 0
        LineNewProratedNP    = (LineNewNP / EAduration) * ProratedDuration if EAduration else 0
        Discount             = self.pricing_engine.get_discount(PricingType, Percent, LineDiscount, UnitCost, UnitList)
        ConsumptionStatus    = self.pricing_engine.get_consumption(TFqty, ExistingQTY, NewQTY, vsADD, vsRETURNED, TIER=TIER)
        LineCM               = self.pricing_engine.get_LineCM(TFqty, vsRETURNED, LineExistingProratedNP)
        LineINV              = self.pricing_engine.get_LineINV(TFqty, LineNewProratedNP)

        row["UnitNP"]                  = UnitNP
        row["LineExistingNP"]          = LineExistingNP
        row["LineNewNP"]               = LineNewNP
        row["LineExistingProratedNP"]  = LineExistingProratedNP
        row["LineNewProratedNP"]       = LineNewProratedNP
        row["Discount"]                = Discount
        row["ConsumptionStatus"]       = ConsumptionStatus
        row["LineCM"]                  = LineCM
        row["LineINV"]                 = LineINV
        return row

    def apply_pricing_logic(self):
        self.rows = [self._process_row(row) for row in self.rows]

    def rename_calculated_columns(self):
        rename = {
            "UnitNP":                 "UNIT NET PRICE",
            "LineExistingNP":         "EXISTNG NET PRICE (months)",
            "LineNewNP":              "NEW NET PRICE (months)",
            "LineExistingProratedNP": "EXISTING PRORATED PRICE",
            "LineNewProratedNP":      "NEW PRORATED PRICE",
            "Discount":               "DISCOUNT OFF LIST",
            "ConsumptionStatus":      "ALLOCATION STATUS",
            "LineCM":                 "PROJECTED CREDIT",
            "LineINV":                "PROJECTED CHARGE"
        }
        self.rows = [{rename.get(k, k): v for k, v in row.items()} for row in self.rows]

    def drop_unwanted_columns(self):
        self.rows = [
            {k.strip(): v for k, v in row.items() if k.strip() not in DROP_COLUMNS}
            for row in self.rows
        ]

    def _compute_delta_qty(self, row):
        existing  = int(row.get("EXISTING QTY", 0) or 0)
        new       = int(row.get("NEW QTY", 0) or 0)
        returned  = int(row.get("(-)LICENSE SHIFT", 0) or 0)
        return 0 if returned == 0 and existing > new else new - existing

    def add_final_metrics(self):
        for row in self.rows:
            delta = self._compute_delta_qty(row)
            row["DELTA QTY"]          = delta
            row["TRUE DELTA NET COST"] = (row.get("PROJECTED CHARGE") or 0) - (row.get("PROJECTED CREDIT") or 0)
            row["TFqty"]              = delta

    def reorder_and_validate(self):
        self.rows = [
            {h: row.get(h) for h in self.final_headers if h in row}
            for row in self.rows
        ]

    def process(self):
        self.normalize_headers()
        self.apply_pricing_logic()
        self.rename_calculated_columns()
        self.drop_unwanted_columns()
        self.add_final_metrics()
        self.reorder_and_validate()
        return self.rows


# ----------------------------  FLASK ROUTE ----------------------------
@truedelta_bp.route('/api/mkdelta', methods=['POST'])
def process_truedelta():
    if 'file' not in request.files: # VALIDATE FILE
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'File must be .xlsx format'}), 400

    user_name    = request.form.get('userName', '').strip() # PARSE FORM FILEDS
    partner_name = request.form.get('partnerName', '').strip()
    partner_acct = request.form.get('partnerAcct', '').strip()
    billing_type = request.form.get('billingType', 'Prepaid').strip()
    pricing_type = request.form.get('pricingType', 'Hold Back').strip()

    try:
        percent_input = float(request.form.get('percent', 0))
        if not (0 <= percent_input <= 100):
            return jsonify({'error': 'Percent must be between 0 and 100'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid percent value'}), 400

    if not all([user_name, partner_name, partner_acct]):
        return jsonify({'error': 'Name, Partner, and Partner Account Number are required'}), 400

    name_parts = user_name.split()
    if len(name_parts) < 2:
        return jsonify({'error': 'Please enter at least a first and last name'}), 400

    first_initial  = name_parts[0][0].upper()
    last_initial   = name_parts[-1][0].upper()
    middle_initial = name_parts[1][0].upper() if len(name_parts) > 2 else ""
    initials   = (first_initial + middle_initial + last_initial).upper()
    date_today = datetime.today().strftime("%Y%m%d")

    user_input = {
        "UserName":     user_name,
        "PartnerName":  partner_name,
        "PartnerAcct":  partner_acct,
        "BillingType":  billing_type,
        "PricingType":  pricing_type,
        "PercentInput": percent_input,
        "Initials":     initials,
        "DateToday":    date_today
    }

    file_bytes    = file.read() # READ UPLOADED FILE WITH openpyxl
    input_buffer  = io.BytesIO(file_bytes)

    try:
        in_wb = load_workbook(input_buffer, data_only=True)
    except Exception:
        return jsonify({'error': 'Could not open Excel file. Make sure it is a valid .xlsx.'}), 400

    if "Summary" not in in_wb.sheetnames:
        return jsonify({'error': 'No "Summary" sheet found in the uploaded file.'}), 400

    summary_ws = in_wb["Summary"]

    sub_id   = str(summary_ws.cell(row=3, column=2).value or "").strip() # sub_id LIVES AT SUMMARY ROW 3 COL B
    tab2_name = sub_id

    if not tab2_name or tab2_name not in in_wb.sheetnames:
        return jsonify({'error': f'Details sheet "{tab2_name}" not found. Expected a sheet named after the subscription reference ID in Summary!B3.'}), 400

    summary_values = [summary_ws.cell(row=i, column=2).value for i in range(1, 11)]

    details_ws   = in_wb[tab2_name]
    details_rows = _sheet_to_dicts(details_ws)

    if not details_rows:
        return jsonify({'error': f'The details sheet "{tab2_name}" appears to be empty.'}), 400

    try: # PROCESS COLUMNS
        processor    = TrueDeltaColumns(
            rows          = details_rows,
            user_input    = user_input,
            pricing_engine= TrueDeltaPricing(),
            header_aliases= SOURCE_HEADER_ALIASES,
            final_headers = DETAIL_TAB_HEADERS
        )
        details_rows = processor.process()
    except Exception as e:
        return jsonify({'error': f'Processing error: {str(e)}'}), 500

    tmp_path = None # BUILD OUTPUT WORKBOOK
    try:
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name

        out_wb = Workbook()
        out_wb.remove(out_wb.active)                  

        out_wb.create_sheet("Summary")

        details_out = out_wb.create_sheet(tab2_name)
        for col_idx, header in enumerate(DETAIL_TAB_HEADERS, 1):
            details_out.cell(row=1, column=col_idx).value = header
        for row_idx, row in enumerate(details_rows, 2):
            for col_idx, header in enumerate(DETAIL_TAB_HEADERS, 1):
                details_out.cell(row=row_idx, column=col_idx).value = row.get(header)

        out_wb.save(tmp_path)

        TrueDeltaDetailsFormat(
            tmp_path, tab2_name,
            total_months=summary_values[8],
            remaining_months=summary_values[9]
        ).format()

        filtered_rows = [r for r in details_rows if (r.get("DELTA QTY") or 0) != 0]

        logo_bytes = None
        if 'logo' in request.files and request.files['logo'].filename:
            logo_bytes = request.files['logo'].read()

        logo_path = os.path.join(os.path.dirname(__file__), '..', 'static', 'LOGO.png')

        TrueDeltaSummaryBuilder(
            filepath      = tmp_path,
            info          = user_input,
            summary_values= summary_values,
            details_rows  = filtered_rows,
            logo_path     = logo_path
        ).build(logo_bytes=logo_bytes)

        TrueDeltaSummaryFormat(tmp_path, "Summary").format(item_end_row=21 + len(filtered_rows))

        base_name = f"TRUE_DELTA_ESTIMATE_{sub_id}_{date_today}_{initials}.xlsx"

        with open(tmp_path, 'rb') as f:
            output_buffer = io.BytesIO(f.read())
        output_buffer.seek(0)

        return send_file(
            output_buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=base_name
        )

    except Exception as e:
        return jsonify({'error': f'Output generation error: {str(e)}'}), 500

    finally:
        if tmp_path and os.path.exists(tmp_path):
            os.unlink(tmp_path)